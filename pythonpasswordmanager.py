#!/usr/bin/env python
# coding: utf-8

# In[72]:


import openpyxl

def openExcelFile(filePath, sheetName):
    #file path = '/Users/kwasidebrah/Downloads/Passwords.xlsx'
    #sheet name = 'Sheet1'
    
    wb = openpyxl.load_workbook(filePath)
    sheet = wb[sheetName]


def displayFullTable(firstEmptyRow):
    
    for i in range(1,firstEmptyRow,1):
        print(i,sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value,sheet.cell(row=i,column=3).value)

        
def findFirstEmptyRow():
    i = 1
    count = 0
    cellValue = 0
    
    while cellValue != None:
        cellValue = sheet.cell(row=i,column=1).value
        count += 1
        i += 1
    
    return count

def addNewPassword(firstEmptyRow):
    accountName = input("What is the account for?")
    username = input("Enter the username: ")
    password = input("Enter the password: ")
    
    sheet.cell(row=firstEmptyRow,column=1).value = accountName
    sheet.cell(row=firstEmptyRow,column=2).value = username
    sheet.cell(row=firstEmptyRow,column=3).value = password
    
def changeExistingRow():
    rowToChange = int(input("Which row would you like to change?"))
    
    accountName = input("What is the new account for?")
    username = input("Enter the new username: ")
    password = input("Enter the new password: ")
    
    sheet.cell(row=rowToChange,column=1).value = accountName
    sheet.cell(row=rowToChange,column=2).value = username
    sheet.cell(row=rowToChange,column=3).value = password
    
def removePasswordFunc():
    rowValue = int(input("Which row do you want removed?"))
    sheet.delete_rows(rowValue)
    
    
#main program starts here:
print("Password Manager using Python and Microsoft Excel")
finished = False



filePath = input("Print out the path to your file.")
sheetName = input("Next, print out the sheets that your passwords are stored in.")

openExcelFile(filePath,sheetName)
firstEmptyRow = findFirstEmptyRow()

while finished != True:
    displayPrompt = input("Would you like to display the full data sheet? (y or n)")

    if displayPrompt == 'y':
        displayFullTable(firstEmptyRow)
    
    addPassword = input("Would you like to add a new password? (y or n)")

    if addPassword == 'y':
        addNewPassword(firstEmptyRow)
        
    changePassword = input("Would you like to change an existing password? (y or n)")
    
    if changePassword == 'y':
        changeExistingRow()
    
    removePassword = input("Would you like to remove a password? (y or n)")

    if removePassword == 'y':
        removePasswordFunc()
    
    finishedOrNot = input("Finished? (y or n)")
    
    if finishedOrNot == 'y':
        finished = True
    else:
        print("Repeating prompts.")
        
    
print("Program ending. Thanks for using!")
    


                            


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




